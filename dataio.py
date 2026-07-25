"""
dataio.py — dataset loading, validation and cleaning
----------------------------------------------------
The scheduler accepts a dataset in three formats:

    JSON   a single .json file (compact, nested)
    CSV    a folder of six .csv files (one table per file)
    XLSX   a single Excel workbook with one tab per table (tidiest for staff)

Real datasets are never clean: fields get left blank, names are typed with
stray spaces or inconsistent capitalisation, rows get duplicated, and numbers
arrive as text. Feeding that straight into a CSP solver produces either a
crash or — worse — a silently wrong timetable.

This module therefore does three jobs, in order:

    1. LOAD       read either format into one internal structure
    2. CLEAN      repair recoverable noise (whitespace, case, numeric text,
                  duplicate rows, unknown-group typos)
    3. VALIDATE   refuse to schedule if anything is still logically broken
                  (missing capacity, group with no size, capacity smaller
                  than every group, unavailability naming nobody, etc.)

Every repair and every refusal is recorded in a report so the user can see
exactly what the system did to their data.

Table layout (same for CSV files and Excel tabs):

    meta          title, task_label, valid_from, valid_to   (dates optional)
    slots         slot, date            (date optional: YYYY-MM-DD)
    venues        name, capacity
    groups        group, size
    tasks         name, staff, groups   (groups pipe-separated: "A|B")
    unavailable   kind, who, slot       (kind: staff | venue)
"""

import csv
import json
import os
import re


# ── report object ────────────────────────────────────────────────────────────

class DataReport:
    """Everything the loader did to (and found wrong with) a dataset."""

    def __init__(self, source, fmt):
        self.source = source
        self.format = fmt
        self.repairs = []   # noise we fixed automatically
        self.errors = []    # problems that make scheduling impossible
        self.warnings = []  # suspicious but survivable

    def repair(self, msg):
        self.repairs.append(msg)

    def error(self, msg):
        self.errors.append(msg)

    def warn(self, msg):
        self.warnings.append(msg)

    @property
    def ok(self):
        return not self.errors

    @property
    def clean(self):
        return not self.errors and not self.repairs and not self.warnings

    def summary(self):
        return (f"{len(self.repairs)} repaired, "
                f"{len(self.warnings)} warning(s), "
                f"{len(self.errors)} error(s)")

    def text(self):
        lines = [f"Dataset: {self.source}  [{self.format.upper()}]"]
        if self.clean:
            lines.append("  No issues found — dataset is clean.")
            return "\n".join(lines)
        for m in self.repairs:
            lines.append(f"  [repaired] {m}")
        for m in self.warnings:
            lines.append(f"  [warning]  {m}")
        for m in self.errors:
            lines.append(f"  [ERROR]    {m}")
        return "\n".join(lines)


# ── small cleaning helpers ───────────────────────────────────────────────────

def _tidy(value):
    """Trim, collapse internal runs of whitespace, drop stray quotes."""
    if value is None:
        return ""
    s = str(value).replace(" ", " ").strip().strip('"').strip("'")
    return re.sub(r"\s+", " ", s)


def _as_int(value):
    """'  60 ' -> 60, '60.0' -> 60, 'sixty' -> None."""
    s = _tidy(value)
    if s == "":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _canon(s):
    """Comparison key that ignores case and spacing: 'Hall  A' == 'hall a'."""
    return _tidy(s).lower()


_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y",
                 "%d %B %Y", "%m/%d/%Y", "%Y/%m/%d")


def _as_date(value):
    """Accept a date as text or as a real Excel date. Returns 'YYYY-MM-DD'."""
    import datetime
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    s = _tidy(value)
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _weekday_of(iso):
    """'2026-08-02' -> 'Sunday'."""
    import datetime
    try:
        return datetime.date.fromisoformat(iso).strftime("%A")
    except (ValueError, TypeError):
        return None


def period_text(data):
    """One line describing when a timetable applies.

    Dated timetables (exams) report the span of their slots; recurring
    timetables (weekly classes) report the term they are valid for.
    Returns "" when the dataset carries no date information at all.
    """
    dates = data.get("slot_dates") or {}
    if dates:
        ordered = [dates[s] for s in data.get("slots", []) if s in dates]
        if ordered:
            lo, hi = min(ordered), max(ordered)
            if lo == hi:
                return pretty_date(lo, "long")
            return f"{pretty_date(lo, 'long')} to {pretty_date(hi, 'long')}"

    lo, hi = data.get("valid_from"), data.get("valid_to")
    if lo and hi:
        return f"Effective {pretty_date(lo, 'long')} to {pretty_date(hi, 'long')}"
    if lo:
        return f"Effective from {pretty_date(lo, 'long')}"
    return ""


def pretty_date(iso, style="short"):
    """'2026-08-02' -> 'Sun 2 Aug 2026' (short) or '2 August 2026' (long)."""
    import datetime
    try:
        d = datetime.date.fromisoformat(iso)
    except (ValueError, TypeError):
        return ""
    if style == "long":
        return f"{d.day} {d.strftime('%B %Y')}"
    return f"{d.strftime('%a')} {d.day} {d.strftime('%b %Y')}"


# ── CSV reading ──────────────────────────────────────────────────────────────

def _read_csv(path):
    """Read a CSV into a list of dicts with tidied, lower-cased headers."""
    if not os.path.exists(path):
        return None
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            rows.append({_canon(k): v for k, v in row.items() if k is not None})
        return rows


# ── Excel reading ────────────────────────────────────────────────────────────

def _read_xlsx(path):
    """Read an .xlsx workbook into {sheet_name: [row dicts]}.

    Row 1 of each sheet is treated as the header. Headers and sheet names are
    canonicalised so 'Venues', 'venues' and ' VENUES ' all resolve the same way.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Reading .xlsx files needs openpyxl.  Install it with:\n"
            "    pip install openpyxl\n"
            "(or use the CSV folder / JSON version of the dataset instead)")

    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets[_canon(ws.title)] = []
            continue
        header = [_canon(h) for h in rows[0]]
        out = []
        for raw in rows[1:]:
            if all(c is None or _tidy(c) == "" for c in raw):
                continue                                   # skip blank rows
            out.append({h: v for h, v in zip(header, raw) if h})
        sheets[_canon(ws.title)] = out
    wb.close()
    return sheets


# ── shared table assembly (used by both CSV and Excel) ───────────────────────

def _build_from_tables(tables, report, label):
    """Assemble the internal structure from {table_name: [row dicts]}.

    Rows are passed through RAW wherever possible: clean() is the single place
    that repairs noise, so JSON, CSV and Excel inputs get identical treatment
    and every repair appears in the report.
    """
    data = {"title": "TIMETABLE", "task_label": "Task",
            "slots": [], "slot_dates": {}, "venues": [], "groups": {},
            "tasks": [], "staff_unavailable": {}, "venue_unavailable": {}}

    def table(name):
        return tables.get(name)

    # meta (optional)
    meta = table("meta")
    if meta:
        row = meta[0]
        data["title"] = _tidy(row.get("title")) or "TIMETABLE"
        data["task_label"] = _tidy(row.get("task_label")) or "Task"
        # a recurring timetable has no per-slot dates, but it does have a
        # period it applies to (a semester, a term)
        for key in ("valid_from", "valid_to"):
            if _tidy(row.get(key)):
                data[key] = row.get(key)
    else:
        report.warn(f"No 'meta' {label} found, using a default title.")

    # slots (an optional 'date' column gives each slot a calendar date)
    rows = table("slots")
    if rows is None:
        report.error(f"The 'slots' {label} is missing.")
    else:
        data["slots"] = [r.get("slot") for r in rows]
        for r in rows:
            slot, date = _tidy(r.get("slot")), r.get("date")
            if slot and date is not None and _tidy(date):
                data["slot_dates"][slot] = date

    # venues
    rows = table("venues")
    if rows is None:
        report.error(f"The 'venues' {label} is missing.")
    else:
        for r in rows:
            data["venues"].append({"name": r.get("name"),
                                   "capacity": r.get("capacity")})

    # groups
    rows = table("groups")
    if rows is None:
        report.error(f"The 'groups' {label} is missing.")
    else:
        for r in rows:
            g = _tidy(r.get("group"))
            if not g:
                report.repair(f"Removed a row with no group name from '{label}' groups.")
                continue
            data["groups"][g] = r.get("size")

    # tasks
    rows = table("tasks")
    if rows is None:
        report.error(f"The 'tasks' {label} is missing.")
    else:
        for r in rows:
            raw_groups = _tidy(r.get("groups"))
            # groups are pipe-separated so commas stay usable inside one cell
            glist = [g for g in re.split(r"[|;]", raw_groups) if _tidy(g)]
            data["tasks"].append({"name": r.get("name"),
                                  "staff": r.get("staff"),
                                  "groups": glist})

    # unavailable (optional)
    rows = table("unavailable")
    if rows:
        for r in rows:
            kind = _canon(r.get("kind"))
            who = _tidy(r.get("who"))
            slot = _tidy(r.get("slot"))
            if not who or not slot:
                continue
            key = "staff_unavailable" if kind.startswith("s") else "venue_unavailable"
            data[key].setdefault(who, []).append(slot)

    return data


def load_csv_folder(folder, report):
    """Assemble the internal data structure from a folder of CSV files."""
    tables = {}
    for name in ("meta", "slots", "venues", "groups", "tasks", "unavailable"):
        rows = _read_csv(os.path.join(folder, f"{name}.csv"))
        if rows is not None:
            tables[name] = rows
    return _build_from_tables(tables, report, "file")


def load_xlsx(path, report):
    """Assemble the internal data structure from an Excel workbook."""
    return _build_from_tables(_read_xlsx(path), report, "sheet")


# ── cleaning ─────────────────────────────────────────────────────────────────

def clean(data, report):
    """Repair recoverable noise in place. Returns the same dict."""

    # --- slots: trim, drop blanks, drop duplicates (keep order) ---
    seen, slots = set(), []
    for s in data.get("slots", []):
        t = _tidy(s)
        if not t:
            report.repair("Removed a blank time slot.")
            continue
        if _canon(t) in seen:
            report.repair(f"Removed duplicate time slot '{t}'.")
            continue
        if t != s:
            report.repair(f"Tidied slot name '{s}' -> '{t}'.")
        seen.add(_canon(t))
        slots.append(t)
    data["slots"] = slots

    # --- slot dates: parse to YYYY-MM-DD, drop dates for unknown slots ---
    dates, valid = {}, {_canon(s): s for s in slots}
    for slot, raw in (data.get("slot_dates") or {}).items():
        s = _tidy(slot)
        if _canon(s) not in valid:
            report.warn(f"Date ignored: '{s}' is not a time slot.")
            continue
        iso = _as_date(raw)
        if iso is None:
            report.warn(f"Date ignored for '{s}': could not read '{raw}'.")
            continue
        dates[valid[_canon(s)]] = iso
    data["slot_dates"] = dates

    # a slot's date should fall on the weekday its name starts with
    for slot, iso in dates.items():
        weekday = _weekday_of(iso)
        first = _tidy(slot).split()[0][:3].lower() if _tidy(slot) else ""
        if weekday and first and not weekday.lower().startswith(first):
            report.warn(f"'{slot}' is dated {iso}, which is a {weekday}.")

    # --- validity period (for recurring timetables) ---
    for key in ("valid_from", "valid_to"):
        if key in data:
            iso = _as_date(data[key])
            if iso is None:
                report.warn(f"Could not read {key.replace('_', ' ')} "
                            f"'{data[key]}', ignoring it.")
                data.pop(key)
            else:
                data[key] = iso
    if data.get("valid_from") and data.get("valid_to") \
            and data["valid_from"] > data["valid_to"]:
        report.warn("Validity period starts after it ends, ignoring both.")
        data.pop("valid_from"), data.pop("valid_to")

    # --- groups: trim names, coerce sizes ---
    groups = {}
    for g, size in data.get("groups", {}).items():
        t = _tidy(g)
        if t != g:
            report.repair(f"Tidied group name '{g}' -> '{t}'.")
        n = _as_int(size)
        if n is None:
            report.error(f"Group '{t}' has no usable size (found: {size!r}).")
            continue
        if n <= 0:
            report.error(f"Group '{t}' has a non-positive size ({n}).")
            continue
        if t in groups:
            report.repair(f"Merged duplicate group entry '{t}'.")
            groups[t] = max(groups[t], n)
        else:
            groups[t] = n
    data["groups"] = groups

    # --- venues: trim, coerce capacity, drop duplicates ---
    venues, vseen = [], set()
    for v in data.get("venues", []):
        raw_name = v.get("name")
        name = _tidy(raw_name)
        if not name:
            report.repair("Removed a venue with no name.")
            continue
        if raw_name is not None and name != str(raw_name):
            report.repair(f"Tidied venue name '{raw_name}' -> '{name}'.")
        cap = _as_int(v.get("capacity"))
        if cap is None:
            report.error(f"Venue '{name}' has no usable capacity "
                         f"(found: {v.get('capacity')!r}).")
            continue
        if cap <= 0:
            report.error(f"Venue '{name}' has a non-positive capacity ({cap}).")
            continue
        if _canon(name) in vseen:
            report.repair(f"Removed duplicate venue '{name}'.")
            continue
        vseen.add(_canon(name))
        venues.append({"name": name, "capacity": cap})
    data["venues"] = venues

    # --- tasks: trim, resolve group typos, drop duplicates ---
    known = {_canon(g): g for g in data["groups"]}
    tasks, tseen = [], set()
    for t in data.get("tasks", []):
        name = _tidy(t.get("name"))
        if not name:
            report.repair("Removed a task with no name.")
            continue
        if _canon(name) in tseen:
            report.repair(f"Removed duplicate task '{name}'.")
            continue
        tseen.add(_canon(name))

        staff = _tidy(t.get("staff"))
        if not staff:
            report.error(f"Task '{name}' has no staff member assigned.")
            continue

        resolved, dropped = [], []
        for g in t.get("groups", []):
            gt = _tidy(g)
            if not gt:
                continue
            if gt in data["groups"]:
                resolved.append(gt)
            elif _canon(gt) in known:                   # case/spacing typo
                fixed = known[_canon(gt)]
                report.repair(f"Task '{name}': group '{gt}' matched to '{fixed}'.")
                resolved.append(fixed)
            else:
                dropped.append(gt)

        if dropped:
            report.error(f"Task '{name}' refers to unknown group(s): "
                         f"{', '.join(dropped)}.")
            continue
        if not resolved:
            report.error(f"Task '{name}' has no student groups attending.")
            continue

        tasks.append({"name": name, "staff": staff,
                      "groups": sorted(set(resolved), key=resolved.index)})
    data["tasks"] = tasks

    # --- unavailability: trim, drop entries naming nobody/nothing ---
    staff_names = {_canon(t["staff"]) for t in data["tasks"]}
    venue_names = {_canon(v["name"]) for v in data["venues"]}
    slot_keys = {_canon(s) for s in data["slots"]}

    for key, universe, label in (
        ("staff_unavailable", staff_names, "staff member"),
        ("venue_unavailable", venue_names, "venue"),
    ):
        cleaned = {}
        for who, slist in (data.get(key) or {}).items():
            w = _tidy(who)
            if _canon(w) not in universe:
                report.warn(f"Unavailability ignored: no {label} named '{w}'.")
                continue
            kept = []
            for s in slist:
                st = _tidy(s)
                if _canon(st) not in slot_keys:
                    report.warn(f"Unavailability ignored: '{w}' listed for "
                                f"unknown slot '{st}'.")
                    continue
                if st not in kept:
                    kept.append(st)
            if kept:
                cleaned[w] = kept
        data[key] = cleaned

    return data


# ── validation ───────────────────────────────────────────────────────────────

def validate(data, report):
    """Logical checks that must pass before scheduling is attempted."""
    if not data["slots"]:
        report.error("No time slots defined.")
    if not data["venues"]:
        report.error("No venues defined.")
    if not data["groups"]:
        report.error("No student groups defined.")
    if not data["tasks"]:
        report.error("No tasks to schedule.")
    if report.errors:
        return data

    biggest_venue = max(v["capacity"] for v in data["venues"])

    # every task must fit somewhere
    for t in data["tasks"]:
        people = sum(data["groups"][g] for g in t["groups"])
        if people > biggest_venue:
            report.error(f"'{t['name']}' needs {people} seats but the largest "
                         f"venue holds {biggest_venue}.")

    # pigeonhole: more tasks than slot x venue cells is impossible
    capacity_cells = len(data["slots"]) * len(data["venues"])
    if len(data["tasks"]) > capacity_cells:
        report.error(f"{len(data['tasks'])} tasks cannot fit into "
                     f"{capacity_cells} slot/venue combinations.")

    # a group cannot have more tasks than there are slots
    for g in data["groups"]:
        n = sum(1 for t in data["tasks"] if g in t["groups"])
        if n > len(data["slots"]):
            report.error(f"Group '{g}' has {n} tasks but only "
                         f"{len(data['slots'])} slots exist.")

    # a staff member cannot run more tasks than there are slots
    load = {}
    for t in data["tasks"]:
        load[t["staff"]] = load.get(t["staff"], 0) + 1
    for staff, n in load.items():
        free = len(data["slots"]) - len(data.get("staff_unavailable", {}).get(staff, []))
        if n > free:
            report.error(f"{staff} has {n} tasks but only {free} available slots.")

    # soft warning: heavy day pressure
    if len(data["tasks"]) > capacity_cells * 0.8:
        report.warn("Dataset is close to full capacity — schedule quality "
                    "may be low even if a valid timetable exists.")

    return data


# ── public entry point ───────────────────────────────────────────────────────

# datasets live in data/ next to the code, but a bare filename is also accepted
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")


def resolve(path):
    """Find a dataset whether it was given as 'x.json' or 'data/x.json'.

    Tries the path as given, then relative to the data/ folder, then relative
    to the project root. Returns the first that exists, else the original path
    so the caller still gets a sensible 'file not found' error.
    """
    root = os.path.dirname(os.path.abspath(__file__))
    for candidate in (path,
                      os.path.join(DATA_DIR, path),
                      os.path.join(root, path)):
        if os.path.exists(candidate):
            return candidate
    return path


def load_dataset(path, strict=True):
    """Load, clean and validate a dataset.

    Accepts a .json file, a folder of .csv files, or an .xlsx workbook.
    Returns (data, report). Raises ValueError if strict and errors were found.
    """
    path = resolve(path)
    name = os.path.basename(path.rstrip("/\\"))

    if os.path.isdir(path):
        report = DataReport(name, "csv")
        data = load_csv_folder(path, report)
    elif path.lower().endswith((".xlsx", ".xlsm")):
        report = DataReport(name, "excel")
        data = load_xlsx(path, report)
    else:
        report = DataReport(name, "json")
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        data.setdefault("staff_unavailable", {})
        data.setdefault("venue_unavailable", {})

    if report.errors:                      # unreadable before cleaning
        if strict:
            raise ValueError(report.text())
        return data, report

    clean(data, report)
    validate(data, report)

    data["_by_name"] = {t["name"]: t for t in data["tasks"]}
    data["_report"] = report

    if strict and not report.ok:
        raise ValueError(report.text())
    return data, report


# ── exporting a JSON dataset to CSV or Excel ────────────────────────────────

def _tables_from_json(json_path):
    """Shared table extraction used by both the CSV and Excel exporters."""
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    unav = []
    for who, slots in (data.get("staff_unavailable") or {}).items():
        unav += [["staff", who, s] for s in slots]
    for who, slots in (data.get("venue_unavailable") or {}).items():
        unav += [["venue", who, s] for s in slots]

    slot_dates = data.get("slot_dates") or {}
    slot_rows = ([[s, slot_dates.get(s, "")] for s in data["slots"]]
                 if slot_dates else [[s] for s in data["slots"]])
    slot_header = ["slot", "date"] if slot_dates else ["slot"]

    meta_header = ["title", "task_label"]
    meta_row = [data.get("title", "TIMETABLE"), data.get("task_label", "Task")]
    if data.get("valid_from") or data.get("valid_to"):
        meta_header += ["valid_from", "valid_to"]
        meta_row += [data.get("valid_from", ""), data.get("valid_to", "")]

    return [
        ("meta", meta_header, [meta_row]),
        ("slots", slot_header, slot_rows),
        ("venues", ["name", "capacity"],
         [[v["name"], v["capacity"]] for v in data["venues"]]),
        ("groups", ["group", "size"],
         [[g, n] for g, n in data["groups"].items()]),
        ("tasks", ["name", "staff", "groups"],
         [[t["name"], t["staff"], "|".join(t["groups"])] for t in data["tasks"]]),
        ("unavailable", ["kind", "who", "slot"], unav),
    ]


def export_to_csv(json_path, folder):
    """Write a JSON dataset out as one CSV file per table."""
    os.makedirs(folder, exist_ok=True)
    for name, header, rows in _tables_from_json(json_path):
        with open(os.path.join(folder, f"{name}.csv"), "w",
                  newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)
    return folder


def export_to_xlsx(json_path, out_path):
    """Write a JSON dataset out as a single Excel workbook, one tab per table."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1A1F36")
    body_font = Font(name="Arial", size=11)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, header, rows in _tables_from_json(json_path):
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for row in rows:
            ws.append(row)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font

        # column widths sized to the longest value in each column
        for i, _name in enumerate(header, start=1):
            longest = max([len(str(header[i - 1]))] +
                          [len(str(r[i - 1])) for r in rows if i <= len(r)] or [0])
            ws.column_dimensions[get_column_letter(i)].width = min(46, longest + 4)

        ws.freeze_panes = "A2"

    # a short legend so a member of staff knows what to edit
    ws = wb.create_sheet("readme", 0)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 92
    guide = [
        ("Sheet", "What to put in it"),
        ("meta", "Title shown above the timetable, and the word for one item (Exam, Class). "
                 "valid_from and valid_to are optional: use them for a repeating "
                 "timetable to show the term it applies to."),
        ("slots", "One time slot per row, e.g. 'Mon 9AM'. The day is the first word. "
                  "The date column is optional: add it to show real calendar dates "
                  "on the timetable. Leave it out for a weekly timetable that repeats."),
        ("venues", "One room per row with its seating capacity as a whole number."),
        ("groups", "One student group per row with how many students are in it."),
        ("tasks", "One exam per row. 'groups' holds the attending groups separated "
                  "by a pipe, e.g. CS-Sem2-A|CS-Sem2-B."),
        ("unavailable", "Optional. kind is 'staff' or 'venue'; who is the name; "
                        "slot must match a row in the slots sheet exactly."),
        ("", ""),
        ("Note", "Extra spaces, capitalisation differences and duplicate rows are "
                 "repaired automatically and reported. Missing capacities, exams with "
                 "no staff member, and unknown group names stop the scheduler."),
    ]
    for row in guide:
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(name="Arial", size=11, bold=True)
        row[1].font = body_font
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    target = sys.argv[1] if len(sys.argv) > 1 else "data/softwarica_exams.json"
    d, rep = load_dataset(target, strict=False)
    print(rep.text())
    print(f"\nLoaded {len(d['tasks'])} tasks, {len(d['venues'])} venues, "
          f"{len(d['groups'])} groups, {len(d['slots'])} slots.")
